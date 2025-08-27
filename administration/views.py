from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count, Avg, Q, Max, Min, Value
from django.db import models
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import io

from authentication.models import User
from exams.models import Course, Exam, Question, UserExam, ExamSession, UserHistory
from .models import Interface, UserInterfaceAccess
from authentication.decorators import admin_required, get_user_interfaces


def is_admin(user):
    """Check if user is an admin."""
    return user.is_authenticated and user.is_admin


@login_required
@user_passes_test(is_admin)
def dashboard_view(request):
    """Admin dashboard with statistics and recent activity."""
    # Calculate statistics
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    total_courses = Course.objects.count()
    active_courses = Course.objects.filter(is_active=True).count()
    total_exams = Exam.objects.count()
    active_exams = Exam.objects.filter(
        start_time__lte=timezone.now(),
        end_time__gte=timezone.now()
    ).count()
    total_submissions = UserExam.objects.filter(exam_sessions__is_submitted=True).distinct().count()
    completed_submissions = total_submissions
    
    # Recent submissions
    recent_submissions = UserExam.objects.filter(
        exam_sessions__is_submitted=True
    ).select_related('user', 'exam').order_by('-exam_sessions__end_time')[:5]
    
    # System alerts (placeholder)
    alerts = []
    
    # Check for exams ending soon
    exams_ending_soon = Exam.objects.filter(
        end_time__gte=timezone.now(),
        end_time__lte=timezone.now() + timedelta(days=1)
    ).count()
    
    if exams_ending_soon > 0:
        alerts.append({
            'type': 'warning',
            'title': 'Exams Ending Soon',
            'message': f'{exams_ending_soon} exam(s) will end within 24 hours.'
        })
    
    # Check for inactive users
    inactive_users = User.objects.filter(
        is_active=True,
        last_login__lt=timezone.now() - timedelta(days=30)
    ).count()
    
    if inactive_users > 0:
        alerts.append({
            'type': 'info',
            'title': 'Inactive Users',
            'message': f'{inactive_users} users haven\'t logged in for over 30 days.'
        })
    
    context = {
        'total_users': total_users,
        'active_users': active_users,
        'total_courses': total_courses,
        'active_courses': active_courses,
        'total_exams': total_exams,
        'active_exams': active_exams,
        'total_submissions': total_submissions,
        'completed_submissions': completed_submissions,
        'recent_submissions': recent_submissions,
        'alerts': alerts,
    }
    
    return render(request, 'administration/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def user_list_view(request):
    """List all users with search and filter options."""
    users = User.objects.all().order_by('-date_joined')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        users = users.filter(
            Q(name__icontains=search_query) |
            Q(national_id__icontains=search_query) |
            Q(position__icontains=search_query)
        )
    
    # Filter by user type (admin/staff/regular)
    user_type = request.GET.get('user_type', '')
    if user_type == 'admin':
        users = users.filter(is_admin=True)
    elif user_type == 'staff':
        users = users.filter(is_staff=True, is_admin=False)
    elif user_type == 'user':
        users = users.filter(is_staff=False, is_admin=False)
    
    # Filter by status
    status = request.GET.get('status', '')
    if status == 'active':
        users = users.filter(is_active=True)
    elif status == 'inactive':
        users = users.filter(is_active=False)
    
    # Pagination
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'user_type': user_type,
        'status': status,
    }
    
    return render(request, 'administration/user_list.html', context)


@login_required
@user_passes_test(is_admin)
def user_create_view(request):
    """Create a new user."""
    if request.method == 'POST':
        # Handle form submission
        name = request.POST.get('name')
        national_id = request.POST.get('national_id')
        position = request.POST.get('position', '')
        is_staff = 'is_staff' in request.POST
        is_admin = 'is_admin' in request.POST
        is_active = 'is_active' in request.POST
        
        # Validate required fields
        if not name or not national_id:
            messages.error(request, 'Name and National ID are required.')
        elif User.objects.filter(national_id=national_id).exists():
            messages.error(request, 'A user with this National ID already exists.')
        else:
            try:
                # Create new user
                user = User.objects.create(
                    name=name,
                    national_id=national_id,
                    position=position,
                    is_staff=is_staff,
                    is_admin=is_admin,
                    is_active=is_active
                )
                messages.success(request, f'User {user.name} created successfully.')
                return redirect('administration:user_list')
            except Exception as e:
                messages.error(request, f'Error creating user: {str(e)}')
    
    context = {
        'title': 'Create User',
        'is_edit': False
    }
    return render(request, 'administration/user_form.html', context)


@login_required
@user_passes_test(is_admin)
def user_edit_view(request, user_id):
    """Edit an existing user."""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        # Handle form submission
        user.name = request.POST.get('name', user.name)
        user.position = request.POST.get('position', user.position)
        user.is_staff = 'is_staff' in request.POST
        user.is_admin = 'is_admin' in request.POST
        user.is_active = 'is_active' in request.POST
        user.save()
        
        messages.success(request, f'User {user.name} updated successfully.')
        return redirect('administration:user_list')
    
    context = {
        'title': 'Edit User',
        'form_user': user,
        'is_edit': True
    }
    return render(request, 'administration/user_form.html', context)


@login_required
@user_passes_test(is_admin)
def user_toggle_status_view(request, user_id):
    """Toggle user active status."""
    user = get_object_or_404(User, id=user_id)
    
    # Prevent admin from deactivating themselves
    if user == request.user:
        messages.error(request, 'You cannot deactivate your own account.')
        return redirect('administration:user_list')
    
    user.is_active = not user.is_active
    user.save()
    
    status = 'activated' if user.is_active else 'deactivated'
    messages.success(request, f'User {user.name} has been {status}.')
    return redirect('administration:user_list')


@login_required
@user_passes_test(is_admin)
def user_delete_view(request, user_id):
    """Delete a user (soft delete by deactivating)."""
    user = get_object_or_404(User, id=user_id)
    
    # Prevent admin from deleting themselves
    if user == request.user:
        messages.error(request, 'You cannot delete your own account.')
        return redirect('administration:user_list')
    
    if request.method == 'POST':
        # Soft delete by deactivating
        user.is_active = False
        user.save()
        messages.success(request, f'User {user.name} has been deleted (deactivated).')
        return redirect('administration:user_list')
    
    # For GET request, show confirmation
    context = {
        'user': user,
        'title': 'Delete User'
    }
    return render(request, 'administration/user_delete_confirm.html', context)


@login_required
@user_passes_test(is_admin)
def course_list_view(request):
    """List all courses."""
    courses = Course.objects.all().order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        courses = courses.filter(
            Q(course_name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(courses, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    
    return render(request, 'administration/course_list.html', context)


@login_required
@user_passes_test(is_admin)
def course_suggestions_api(request):
    """API endpoint for course name suggestions."""
    query = request.GET.get('q', '').strip()
    suggestions = []
    
    if query and len(query) >= 2:
        # Get existing course names that match the query
        courses = Course.objects.filter(
            course_name__icontains=query
        ).values_list('course_name', flat=True).distinct()[:10]
        
        suggestions = list(courses)
        
        # Add some common course name patterns if no exact matches
        if not suggestions:
            common_patterns = [
                f"{query} Fundamentals",
                f"Introduction to {query}",
                f"Advanced {query}",
                f"{query} Basics",
                f"{query} Course"
            ]
            suggestions = common_patterns[:5]
    
    return JsonResponse({'suggestions': suggestions})


@login_required
@user_passes_test(is_admin)
def course_create_view(request):
    """Create a new course."""
    if request.method == 'POST':
        course_name = request.POST.get('name')
        description = request.POST.get('description')
        
        if course_name:
            try:
                course = Course.objects.create(
                    course_name=course_name,
                    description=description or ''
                )
                messages.success(request, f'Course "{course_name}" created successfully!')
                return redirect('administration:course_list')
            except Exception as e:
                messages.error(request, f'Error creating course: {str(e)}')
        else:
            messages.error(request, 'Course name is required.')
    
    context = {'title': 'Create Course'}
    return render(request, 'administration/course_form.html', context)


@login_required
@user_passes_test(is_admin)
def course_edit_view(request, course_id):
    """Edit an existing course."""
    course = get_object_or_404(Course, course_id=course_id)
    
    if request.method == 'POST':
        course_name = request.POST.get('name')
        description = request.POST.get('description')
        
        if course_name:
            try:
                course.course_name = course_name
                course.description = description or ''
                course.save()
                messages.success(request, f'Course "{course_name}" updated successfully!')
                return redirect('administration:course_list')
            except Exception as e:
                messages.error(request, f'Error updating course: {str(e)}')
        else:
            messages.error(request, 'Course name is required.')
    
    context = {
        'title': 'Edit Course',
        'course': course,
        'is_edit': True
    }
    return render(request, 'administration/course_form.html', context)


@login_required
@user_passes_test(is_admin)
def course_delete_view(request, course_id):
    """Delete a course."""
    course = get_object_or_404(Course, course_id=course_id)
    
    if request.method == 'POST':
        course_name = course.course_name
        course.delete()
        messages.success(request, f'Course "{course_name}" has been deleted successfully.')
        return redirect('administration:course_list')
    
    # For GET request, show confirmation
    context = {
        'course': course,
        'title': 'Delete Course'
    }
    return render(request, 'administration/course_delete_confirm.html', context)


@login_required
@user_passes_test(is_admin)
def exam_list_view(request):
    """List all exams."""
    exams = Exam.objects.select_related('course').order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        exams = exams.filter(
            Q(exam_title__icontains=search_query) |
            Q(course__course_name__icontains=search_query)
        )
    
    # Filter by course
    course_id = request.GET.get('course', '')
    if course_id:
        exams = exams.filter(course_id=course_id)
    
    # Pagination
    paginator = Paginator(exams, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    courses = Course.objects.all().order_by('course_name')
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'course_id': course_id,
        'courses': courses,
    }
    
    return render(request, 'administration/exam_list.html', context)


@login_required
@user_passes_test(is_admin)
def exam_create_view(request):
    """Create a new exam."""
    if request.method == 'POST':
        exam_title = request.POST.get('exam_title')
        course_id = request.POST.get('course')
        time_limit = request.POST.get('time_limit')
        description = request.POST.get('description')
        total_points = request.POST.get('total_points')
        passing_score = request.POST.get('passing_score')
        max_attempts = request.POST.get('max_attempts')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        
        if exam_title and course_id and time_limit:
            try:
                course = Course.objects.get(course_id=course_id)
                exam = Exam.objects.create(
                    exam_title=exam_title,
                    course=course,
                    time_limit=int(time_limit),
                    description=description or '',
                    total_points=float(total_points) if total_points else None,
                    passing_score=int(passing_score) if passing_score else 60,
                    max_attempts=int(max_attempts) if max_attempts else 1,
                    start_time=start_time if start_time else None,
                    end_time=end_time if end_time else None
                )
                messages.success(request, f'Exam "{exam_title}" created successfully!')
                return redirect('administration:exam_list')
            except Course.DoesNotExist:
                messages.error(request, 'Selected course does not exist.')
            except ValueError:
                messages.error(request, 'Time limit must be a valid number.')
            except Exception as e:
                messages.error(request, f'Error creating exam: {str(e)}')
        else:
            messages.error(request, 'All fields are required.')
    
    courses = Course.objects.all().order_by('course_name')
    context = {
        'title': 'Create Exam',
        'courses': courses
    }
    return render(request, 'administration/exam_form.html', context)


@login_required
@user_passes_test(is_admin)
def exam_edit_view(request, exam_id):
    """Edit an existing exam."""
    exam = get_object_or_404(Exam, exam_id=exam_id)
    
    if request.method == 'POST':
        exam_title = request.POST.get('exam_title')
        course_id = request.POST.get('course')
        time_limit = request.POST.get('time_limit')
        description = request.POST.get('description')
        total_points = request.POST.get('total_points')
        passing_score = request.POST.get('passing_score')
        max_attempts = request.POST.get('max_attempts')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        
        if exam_title and course_id and time_limit:
            try:
                course = Course.objects.get(course_id=course_id)
                exam.exam_title = exam_title
                exam.course = course
                exam.time_limit = int(time_limit)
                exam.description = description or ''
                exam.total_points = float(total_points) if total_points else None
                exam.passing_score = int(passing_score) if passing_score else 60
                exam.max_attempts = int(max_attempts) if max_attempts else 1
                
                # Handle datetime fields
                if start_time:
                    exam.start_time = start_time
                if end_time:
                    exam.end_time = end_time
                    
                exam.save()
                messages.success(request, f'Exam "{exam_title}" updated successfully!')
                return redirect('administration:exam_list')
            except Course.DoesNotExist:
                messages.error(request, 'Selected course does not exist.')
            except ValueError:
                messages.error(request, 'Time limit must be a valid number.')
            except Exception as e:
                messages.error(request, f'Error updating exam: {str(e)}')
        else:
            messages.error(request, 'Title, course, and time limit are required.')
    
    courses = Course.objects.all().order_by('course_name')
    context = {
        'title': 'Edit Exam',
        'exam': exam,
        'courses': courses,
        'is_edit': True
    }
    return render(request, 'administration/exam_form.html', context)


@login_required
@user_passes_test(is_admin)
def exam_delete_view(request, exam_id):
    """Delete an exam."""
    exam = get_object_or_404(Exam, exam_id=exam_id)
    
    if request.method == 'POST':
        exam_title = exam.exam_title
        exam.delete()
        messages.success(request, f'Exam "{exam_title}" has been deleted successfully.')
        return redirect('administration:exam_list')
    
    # For GET request, show confirmation
    context = {
        'exam': exam,
        'title': 'Delete Exam'
    }
    return render(request, 'administration/exam_delete_confirm.html', context)


@login_required
@user_passes_test(is_admin)
def question_list_view(request):
    """List all questions."""
    questions = Question.objects.select_related('exam').order_by('-question_id')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        questions = questions.filter(
            Q(question_text__icontains=search_query) |
            Q(exam__exam_title__icontains=search_query)
        )
    
    # Filter by exam
    exam_id = request.GET.get('exam', '')
    if exam_id:
        questions = questions.filter(exam_id=exam_id)
    
    # Filter by question type
    question_type = request.GET.get('type', '')
    if question_type:
        questions = questions.filter(question_type=question_type)
    
    # Pagination
    paginator = Paginator(questions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    exams = Exam.objects.all().order_by('exam_title')
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'exam_id': exam_id,
        'question_type': question_type,
        'exams': exams,
        'question_types': Question.QUESTION_TYPES,
    }
    
    return render(request, 'administration/question_list.html', context)


@login_required
@user_passes_test(is_admin)
def question_create_view(request):
    """Create a new question."""
    if request.method == 'POST':
        exam_id = request.POST.get('exam')
        question_type = request.POST.get('question_type')
        question_text = request.POST.get('question_text')
        
        if exam_id and question_type and question_text:
            try:
                exam = Exam.objects.get(exam_id=exam_id)
                points = request.POST.get('points', 1.0)
                question = Question.objects.create(
                    exam=exam,
                    question_type=question_type,
                    question_text=question_text,
                    points=float(points) if points else 1.0
                )
                
                # Handle different question types
                if question_type == 'multiple_choice':
                    question.option_a = request.POST.get('option_a', '')
                    question.option_b = request.POST.get('option_b', '')
                    question.option_c = request.POST.get('option_c', '')
                    question.option_d = request.POST.get('option_d', '')
                    question.correct_answer = request.POST.get('correct_answer')
                    question.save()
                    
                elif question_type == 'true_false':
                    question.option_a = 'True'
                    question.option_b = 'False'
                    question.correct_answer = request.POST.get('correct_answer')
                    question.save()
                    
                else:
                    question.correct_answer = request.POST.get('correct_answer', '')
                    question.save()
                
                messages.success(request, f'Question created successfully!')
                return redirect('administration:question_list')
                
            except Exam.DoesNotExist:
                messages.error(request, 'Selected exam does not exist.')

            except Exception as e:
                messages.error(request, f'Error creating question: {str(e)}')
        else:
            messages.error(request, 'All required fields must be filled.')
    
    exams = Exam.objects.select_related('course').order_by('exam_title')
    context = {
        'title': 'Create Question',
        'exams': exams
    }
    return render(request, 'administration/question_form.html', context)


@login_required
@user_passes_test(is_admin)
def assignment_create_view(request):
    """Assign exams to users."""
    if request.method == 'POST':
        user_id = request.POST.get('user')
        exam_id = request.POST.get('exam')
        due_date = request.POST.get('due_date')
        attempts_allowed = request.POST.get('attempts_allowed')
        
        # Validate required fields
        if not user_id or not exam_id:
            messages.error(request, 'Please select both a student and an exam.')
        else:
            try:
                user = User.objects.get(id=user_id)
                exam = Exam.objects.get(exam_id=exam_id)
                
                # Check if assignment already exists
                if UserExam.objects.filter(user=user, exam=exam).exists():
                    messages.error(request, f'Assignment already exists for {user.name} and {exam.exam_title}.')
                else:
                    # Create the assignment
                    assignment = UserExam.objects.create(
                        user=user,
                        exam=exam,
                        due_date=due_date if due_date else None,
                        attempts_allowed=int(attempts_allowed) if attempts_allowed else None,
                        assigned_date=timezone.now()
                    )
                    messages.success(request, f'Assignment created successfully for {user.name}.')
                    return redirect('administration:assignment_list')
            except (User.DoesNotExist, Exam.DoesNotExist):
                messages.error(request, 'Invalid user or exam selected.')
            except ValueError:
                messages.error(request, 'Invalid attempts allowed value.')
    
    # Get users and exams for the form
    users = User.objects.filter(is_active=True).order_by('name')
    exams = Exam.objects.select_related('course').order_by('exam_title')
    
    context = {
        'title': 'Create Assignment',
        'users': users,
        'exams': exams,
    }
    return render(request, 'administration/assignment_form.html', context)


@login_required
@user_passes_test(is_admin)
def assignment_list_view(request):
    """List all exam assignments."""
    assignments = UserExam.objects.select_related('user', 'exam', 'exam__course').prefetch_related('exam_sessions').order_by('-assigned_date')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        assignments = assignments.filter(
            Q(user__name__icontains=search_query) |
            Q(user__national_id__icontains=search_query) |
            Q(exam__exam_title__icontains=search_query)
        )
    
    # Filter by exam
    exam_id = request.GET.get('exam', '')
    if exam_id:
        assignments = assignments.filter(exam_id=exam_id)
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        if status_filter == 'completed':
            # Filter assignments that have submitted exam sessions
            assignments = assignments.filter(exam_sessions__is_submitted=True).distinct()
        elif status_filter == 'pending':
            # Filter assignments that don't have submitted exam sessions
            assignments = assignments.exclude(exam_sessions__is_submitted=True)
    
    # Add assignment_status annotation to assignments
    from django.db.models import Case, When, Exists, OuterRef
    
    assignments = assignments.annotate(
        assignment_status=Case(
            When(Exists(ExamSession.objects.filter(user_exam=OuterRef('pk'), is_submitted=True)), then=Value('completed')),
            default=Value('pending'),
            output_field=models.CharField()
        )
    )
    
    # Pagination
    paginator = Paginator(assignments, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    exams = Exam.objects.all().order_by('exam_title')
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'exam_id': exam_id,
        'status_filter': status_filter,
        'exams': exams,
    }
    
    return render(request, 'administration/assignment_list.html', context)


@login_required
@user_passes_test(is_admin)
def reports_view(request):
    """Generate and display various reports."""
    # Exam performance statistics
    exam_stats = Exam.objects.annotate(
        total_attempts=Count('user_assignments'),
        completed_attempts=Count('user_assignments', filter=Q(user_assignments__exam_sessions__is_submitted=True)),
        avg_score=Value(0.0)  # Placeholder - score calculation would need to be implemented
    ).order_by('-created_at')[:10]
    
    # User performance statistics
    user_stats = User.objects.annotate(
        total_exams=Count('assigned_exams'),
        completed_exams=Count('assigned_exams', filter=Q(assigned_exams__exam_sessions__is_submitted=True)),
        avg_score=Value(0.0)  # Placeholder - score calculation would need to be implemented
    ).filter(is_admin=False).order_by('-total_exams')[:10]
    
    # Recent activity
    recent_sessions = ExamSession.objects.select_related(
        'user_exam__user', 'user_exam__exam'
    ).order_by('-start_time')[:20]
    
    context = {
        'exam_stats': exam_stats,
        'user_stats': user_stats,
        'recent_sessions': recent_sessions,
    }
    
    return render(request, 'administration/reports.html', context)


@login_required
@user_passes_test(is_admin)
def user_history_view(request):
    """View user activity history."""
    # This would typically show login history, exam attempts, etc.
    # For now, we'll show exam sessions
    sessions = ExamSession.objects.select_related(
        'user_exam__user', 'user_exam__exam'
    ).order_by('-start_time')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        sessions = sessions.filter(
            Q(user_exam__user__name__icontains=search_query) |
            Q(user_exam__user__national_id__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(sessions, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    
    return render(request, 'administration/user_history.html', context)


@login_required
@user_passes_test(is_admin)
def exam_results_view(request, exam_id):
    """View results for a specific exam."""
    exam = get_object_or_404(Exam, exam_id=exam_id)
    
    # Get all user exams for this exam
    user_exams = UserExam.objects.filter(exam=exam).select_related('user')
    
    # Filter by status (based on exam sessions)
    status_filter = request.GET.get('status', '')
    if status_filter:
        if status_filter == 'completed':
            user_exams = user_exams.filter(exam_sessions__is_submitted=True).distinct()
        elif status_filter == 'in_progress':
            user_exams = user_exams.filter(exam_sessions__is_submitted=False).distinct()
        elif status_filter == 'not_started':
            user_exams = user_exams.exclude(exam_sessions__isnull=False)
    
    # Filter by score range
    min_score = request.GET.get('min_score', '')
    max_score = request.GET.get('max_score', '')
    if min_score:
        try:
            user_exams = user_exams.filter(score__gte=float(min_score))
        except ValueError:
            pass
    if max_score:
        try:
            user_exams = user_exams.filter(score__lte=float(max_score))
        except ValueError:
            pass
    
    # Search by user name or national ID
    search_query = request.GET.get('search', '')
    if search_query:
        user_exams = user_exams.filter(
            Q(user__name__icontains=search_query) |
            Q(user__national_id__icontains=search_query)
        )
    
    # Calculate statistics
    total_attempts = user_exams.count()
    completed_attempts = user_exams.filter(exam_sessions__is_submitted=True).distinct().count()
    in_progress_attempts = user_exams.filter(exam_sessions__is_submitted=False).distinct().count()
    not_started_attempts = user_exams.exclude(exam_sessions__isnull=False).count()
    
    # Get completed exams with scores from exam sessions
    completed_sessions = ExamSession.objects.filter(
        user_exam__in=user_exams,
        is_submitted=True
    ).select_related('user_exam')
    # Calculate average score from completed sessions
    # Note: Score calculation would need to be implemented in ExamSession model
    # For now, we'll set avg_score to 0 until score calculation is implemented
    avg_score = 0
    
    # Calculate pass rate (assuming 60% is passing)
    passing_score = 60  # Default passing score
    passed_count = 0  # Will need to implement score calculation
    pass_rate = 0  # Will need to implement score calculation
    
    # Pagination
    paginator = Paginator(user_exams.order_by('-assigned_date'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'exam': exam,
        'page_obj': page_obj,
        'total_attempts': total_attempts,
        'completed_attempts': completed_attempts,
        'in_progress_attempts': in_progress_attempts,
        'not_started_attempts': not_started_attempts,
        'avg_score': avg_score,
        'pass_rate': pass_rate,
        'search_query': search_query,
        'status_filter': status_filter,
        'min_score': min_score,
        'max_score': max_score,
    }
    
    return render(request, 'administration/exam_results.html', context)


@login_required
@user_passes_test(is_admin)
def user_exam_detail_view(request, user_exam_id):
    """View detailed results for a specific user's exam attempt."""
    user_exam = get_object_or_404(UserExam, user_exam_id=user_exam_id)
    
    # Get the exam session
    try:
        session = ExamSession.objects.get(user_exam=user_exam)
    except ExamSession.DoesNotExist:
        session = None
    
    # Get user's answers
    user_answers = []
    if session:
        answers = UserHistory.objects.filter(session=session).select_related('question')
        for answer in answers:
            is_correct = False
            points_earned = 0
            
            if answer.question.question_type == 'multiple_choice':
                is_correct = answer.selected_answer == answer.question.correct_answer
                points_earned = answer.question.points if is_correct else 0
            elif answer.question.question_type == 'true_false':
                is_correct = answer.selected_answer == answer.question.correct_answer
                points_earned = answer.question.points if is_correct else 0
            # For short_answer and essay, manual grading would be needed
            
            user_answers.append({
                'question': answer.question,
                'user_answer': answer.selected_answer or answer.answer_text,
                'correct_answer': answer.question.correct_answer,
                'is_correct': is_correct,
                'points_earned': points_earned,
                'answer_obj': answer
            })
    
    # Calculate detailed statistics
    total_questions = user_exam.exam.questions.count()
    answered_questions = len([a for a in user_answers if a['user_answer']])
    correct_answers = len([a for a in user_answers if a['is_correct']])
    total_points = sum(q.points for q in user_exam.exam.questions.all())
    earned_points = sum(a['points_earned'] for a in user_answers)
    
    context = {
        'user_exam': user_exam,
        'session': session,
        'user_answers': user_answers,
        'total_questions': total_questions,
        'answered_questions': answered_questions,
        'correct_answers': correct_answers,
        'total_points': total_points,
        'earned_points': earned_points,
    }
    
    return render(request, 'administration/user_exam_detail.html', context)


@login_required
@user_passes_test(is_admin)
def individual_user_history_view(request, user_id):
    """View detailed history for a specific user."""
    user = get_object_or_404(User, id=user_id)
    
    # Get user's exam history
    user_exams = UserExam.objects.filter(user=user).select_related('exam', 'exam__course').order_by('-assigned_date')
    
    # Get user's exam sessions
    sessions = ExamSession.objects.filter(user_exam__user=user).select_related('user_exam__exam').order_by('-start_time')
    
    # Calculate user statistics
    total_assigned = user_exams.count()
    total_completed = user_exams.filter(exam_sessions__is_submitted=True).distinct().count()
    total_in_progress = user_exams.filter(exam_sessions__isnull=False, exam_sessions__is_submitted=False).distinct().count()
    total_not_started = total_assigned - total_completed - total_in_progress
    
    # Note: Score calculation would need to be implemented and stored
    # For now, setting scores to 0 until score storage is implemented
    avg_score = 0
    highest_score = 0
    lowest_score = 0
    
    # Recent activity (last 10 sessions)
    recent_sessions = sessions[:10]
    
    context = {
        'user': user,
        'user_exams': user_exams,
        'recent_sessions': recent_sessions,
        'total_assigned': total_assigned,
        'total_completed': total_completed,
        'total_in_progress': total_in_progress,
        'total_not_started': total_not_started,
        'avg_score': avg_score,
        'highest_score': highest_score,
        'lowest_score': lowest_score,
    }
    
    return render(request, 'administration/user_history.html', context)


@login_required
@admin_required
def interface_list_view(request):
    """Display list of all interfaces"""
    interfaces = Interface.objects.all().order_by('module_name', 'function')
    
    return render(request, 'administration/interface_list.html', {
        'interfaces': interfaces,
    })


@login_required
@admin_required
def interface_create_view(request):
    """Create a new interface"""
    if request.method == 'POST':
        module_name = request.POST.get('module_name')
        function = request.POST.get('function')
        url = request.POST.get('url')
        
        if module_name and function and url:
            try:
                interface = Interface.objects.create(
                    module_name=module_name,
                    function=function,
                    url=url
                )
                messages.success(request, f'Interface "{interface.module_name} - {interface.function}" created successfully.')
                return redirect('administration:interface_list')
            except Exception as e:
                messages.error(request, f'Error creating interface: {str(e)}')
        else:
            messages.error(request, 'All fields are required.')
    
    return render(request, 'administration/interface_form.html')


@login_required
@admin_required
def user_interface_access_view(request, user_id):
    """Manage interface access for a specific user"""
    user = get_object_or_404(User, pk=user_id)
    interfaces = Interface.objects.all().order_by('module_name', 'function')
    
    # Get current user access
    user_access = {}
    for access in UserInterfaceAccess.objects.filter(user=user):
        user_access[access.interface.interface_id] = access.has_access
    
    if request.method == 'POST':
        # Update interface access
        selected_interfaces = request.POST.getlist('interfaces')
        
        # Remove all existing access for this user
        UserInterfaceAccess.objects.filter(user=user).delete()
        
        # Add new access records
        for interface_id in selected_interfaces:
            try:
                interface = Interface.objects.get(interface_id=interface_id)
                UserInterfaceAccess.objects.create(
                    user=user,
                    interface=interface,
                    has_access=True,
                    granted_by=request.user
                )
            except Interface.DoesNotExist:
                continue
        
        messages.success(request, f'Interface access updated for {user.name}.')
        return redirect('administration:user_list')
    
    return render(request, 'administration/user_interface_access.html', {
        'user_obj': user,
        'interfaces': interfaces,
        'user_access': user_access,
    })


@login_required
@admin_required
def bulk_interface_access_view(request):
    """Manage interface access for multiple users"""
    users = User.objects.filter(is_active=True).order_by('name')
    interfaces = Interface.objects.all().order_by('module_name', 'function')
    
    if request.method == 'POST':
        selected_users = request.POST.getlist('users')
        selected_interfaces = request.POST.getlist('interfaces')
        action = request.POST.get('action')  # 'grant' or 'revoke'
        
        if selected_users and selected_interfaces and action:
            updated_count = 0
            
            for user_id in selected_users:
                try:
                    user = User.objects.get(pk=user_id)
                    
                    for interface_id in selected_interfaces:
                        try:
                            interface = Interface.objects.get(interface_id=interface_id)
                            
                            # Get or create access record
                            access, created = UserInterfaceAccess.objects.get_or_create(
                                user=user,
                                interface=interface,
                                defaults={
                                    'has_access': action == 'grant',
                                    'granted_by': request.user
                                }
                            )
                            
                            if not created:
                                access.has_access = action == 'grant'
                                access.granted_by = request.user
                                access.granted_date = timezone.now()
                                access.save()
                            
                            updated_count += 1
                            
                        except Interface.DoesNotExist:
                            continue
                            
                except User.DoesNotExist:
                    continue
            
            action_text = 'granted' if action == 'grant' else 'revoked'
            messages.success(request, f'Interface access {action_text} for {len(selected_users)} users across {len(selected_interfaces)} interfaces.')
            
        else:
            messages.error(request, 'Please select users, interfaces, and an action.')
    
    return render(request, 'administration/bulk_interface_access.html', {
        'users': users,
        'interfaces': interfaces,
    })


@login_required
@user_passes_test(is_admin)
def export_user_history_excel(request, user_id):
    """Export user history to Excel format"""
    user = get_object_or_404(User, user_id=user_id)
    
    # Get user exam history
    user_exams = UserExam.objects.filter(user=user).select_related('exam', 'exam__course').order_by('-assigned_date')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"User History - {user.full_name}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"Exam History Report - {user.full_name}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Add user info
    ws['A3'] = "National ID:"
    ws['B3'] = user.national_id
    ws['A4'] = "Email:"
    ws['B4'] = user.email
    ws['A5'] = "Generated:"
    ws['B5'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Add headers
    headers = ['Exam Name', 'Course', 'Status', 'Assigned Date', 'Due Date', 'Attempts', 'Score']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row, user_exam in enumerate(user_exams, 8):
        ws.cell(row=row, column=1, value=user_exam.exam.title)
        ws.cell(row=row, column=2, value=user_exam.exam.course.name)
        ws.cell(row=row, column=3, value=user_exam.status.title())
        ws.cell(row=row, column=4, value=user_exam.assigned_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=5, value=user_exam.due_date.strftime('%Y-%m-%d') if user_exam.due_date else 'No due date')
        ws.cell(row=row, column=6, value=user_exam.get_attempts_taken())
        
        # Get best score
        best_session = ExamSession.objects.filter(
            user_exam=user_exam, 
            submitted_at__isnull=False
        ).order_by('-score').first()
        score = f"{best_session.score}%" if best_session and best_session.score is not None else "Not taken"
        ws.cell(row=row, column=7, value=score)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="user_history_{user.national_id}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def export_user_history_pdf(request, user_id):
    """Export user history to PDF format"""
    user = get_object_or_404(User, user_id=user_id)
    
    # Get user exam history
    user_exams = UserExam.objects.filter(user=user).select_related('exam', 'exam__course').order_by('-assigned_date')
    
    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Add title
    title = Paragraph(f"Exam History Report - {user.full_name}", title_style)
    elements.append(title)
    
    # Add user info
    user_info = [
        ["National ID:", user.national_id],
        ["Email:", user.email],
        ["Generated:", timezone.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    user_info_table = Table(user_info, colWidths=[2*inch, 4*inch])
    user_info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(user_info_table)
    elements.append(Spacer(1, 20))
    
    # Prepare exam data
    data = [['Exam Name', 'Course', 'Status', 'Assigned', 'Due Date', 'Attempts', 'Score']]
    
    for user_exam in user_exams:
        # Get best score
        best_session = ExamSession.objects.filter(
            user_exam=user_exam, 
            submitted_at__isnull=False
        ).order_by('-score').first()
        score = f"{best_session.score}%" if best_session and best_session.score is not None else "Not taken"
        
        data.append([
            user_exam.exam.title[:20] + '...' if len(user_exam.exam.title) > 20 else user_exam.exam.title,
            user_exam.exam.course.name[:15] + '...' if len(user_exam.exam.course.name) > 15 else user_exam.exam.course.name,
            user_exam.status.title(),
            user_exam.assigned_date.strftime('%Y-%m-%d'),
            user_exam.due_date.strftime('%Y-%m-%d') if user_exam.due_date else 'No due date',
            str(user_exam.get_attempts_taken()),
            score
        ])
    
    # Create table
    table = Table(data, colWidths=[1.5*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.7*inch])
    table.setStyle(TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Body style
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="user_history_{user.national_id}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    
    return response


@login_required
@user_passes_test(is_admin)
def export_reports_excel(request):
    """Export reports data to Excel format"""
    
    # Get exam performance data
    exams = Exam.objects.annotate(
        total_assigned=Count('user_assignments'),
        total_completed=Count('user_assignments__exam_sessions', filter=Q(user_assignments__exam_sessions__is_submitted=True))
        # Note: Score is calculated but not stored in ExamSession model
    ).select_related('course')
    
    # Get user performance data
    users = User.objects.annotate(
        total_assigned=Count('assigned_exams'),
        total_completed=Count('assigned_exams__exam_sessions', filter=Q(assigned_exams__exam_sessions__is_submitted=True))
        # Note: Score is calculated but not stored in ExamSession model
    )
    
    # Get recent activity
    recent_sessions = ExamSession.objects.filter(
        is_submitted=True
    ).select_related('user_exam__user', 'user_exam__exam').order_by('-end_time')[:20]
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Sheet 1: Exam Performance
    ws1 = wb.active
    ws1.title = "Exam Performance"
    
    # Add title
    ws1.merge_cells('A1:F1')
    title_cell = ws1['A1']
    title_cell.value = "Exam Performance Report"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Add generation date
    ws1['A3'] = "Generated:"
    ws1['B3'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Add headers
    exam_headers = ['Exam Name', 'Course', 'Total Assigned', 'Total Completed', 'Completion Rate', 'Average Score']
    for col, header in enumerate(exam_headers, 1):
        cell = ws1.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add exam data
    for row, exam in enumerate(exams, 6):
        ws1.cell(row=row, column=1, value=exam.exam_title)
        ws1.cell(row=row, column=2, value=exam.course.course_name)
        ws1.cell(row=row, column=3, value=exam.total_assigned or 0)
        ws1.cell(row=row, column=4, value=exam.total_completed or 0)
        completion_rate = (exam.total_completed / exam.total_assigned * 100) if exam.total_assigned > 0 else 0
        ws1.cell(row=row, column=5, value=f"{completion_rate:.1f}%")
        # Since score is not stored in ExamSession, we can't calculate average score here
        ws1.cell(row=row, column=6, value="N/A")
    
    # Sheet 2: User Performance
    ws2 = wb.create_sheet("User Performance")
    
    # Add title
    ws2.merge_cells('A1:F1')
    title_cell = ws2['A1']
    title_cell.value = "User Performance Report"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Add generation date
    ws2['A3'] = "Generated:"
    ws2['B3'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Add headers
    user_headers = ['Full Name', 'National ID', 'Total Assigned', 'Total Completed', 'Completion Rate', 'Average Score']
    for col, header in enumerate(user_headers, 1):
        cell = ws2.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add user data
    for row, user in enumerate(users, 6):
        ws2.cell(row=row, column=1, value=user.name)
        ws2.cell(row=row, column=2, value=user.national_id)
        ws2.cell(row=row, column=3, value=user.total_assigned or 0)
        ws2.cell(row=row, column=4, value=user.total_completed or 0)
        completion_rate = (user.total_completed / user.total_assigned * 100) if user.total_assigned > 0 else 0
        ws2.cell(row=row, column=5, value=f"{completion_rate:.1f}%")
        # Since score is not stored in ExamSession, we can't calculate average score here
        ws2.cell(row=row, column=6, value="N/A")
    
    # Sheet 3: Recent Activity
    ws3 = wb.create_sheet("Recent Activity")
    
    # Add title
    ws3.merge_cells('A1:E1')
    title_cell = ws3['A1']
    title_cell.value = "Recent Activity Report"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Add generation date
    ws3['A3'] = "Generated:"
    ws3['B3'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Add headers
    activity_headers = ['User', 'Exam', 'Score', 'Submitted Date', 'Duration']
    for col, header in enumerate(activity_headers, 1):
        cell = ws3.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add activity data
    for row, session in enumerate(recent_sessions, 6):
        ws3.cell(row=row, column=1, value=session.user_exam.user.name)
        ws3.cell(row=row, column=2, value=session.user_exam.exam.exam_title)
        ws3.cell(row=row, column=3, value="N/A")  # Score is calculated but not stored in ExamSession
        ws3.cell(row=row, column=4, value=session.end_time.strftime('%Y-%m-%d %H:%M') if session.end_time else "N/A")
        
        # Calculate duration
        if session.start_time and session.end_time:
            duration = session.end_time - session.start_time
            hours, remainder = divmod(duration.total_seconds(), 3600)
            minutes, _ = divmod(remainder, 60)
            duration_str = f"{int(hours)}h {int(minutes)}m"
        else:
            duration_str = "N/A"
        ws3.cell(row=row, column=5, value=duration_str)
    
    # Auto-adjust column widths for all sheets
    for ws in [ws1, ws2, ws3]:
        for column in ws.columns:
            max_length = 0
            # Skip empty columns
            if not column:
                continue
                
            # Handle potential merged cells
            try:
                column_letter = column[0].column_letter
            except AttributeError:
                # Skip cells without column_letter attribute (like merged cells)
                continue
                
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reports_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def export_reports_pdf(request):
    """Export reports data to PDF format"""
    
    # Get exam performance data
    exams = Exam.objects.annotate(
        total_assigned=Count('user_assignments'),
        total_completed=Count('user_assignments__exam_sessions', filter=Q(user_assignments__exam_sessions__is_submitted=True))
        # Note: Score is calculated but not stored in ExamSession model
    ).select_related('course')[:10]  # Limit for PDF
    
    # Get user performance data
    users = User.objects.annotate(
        total_assigned=Count('assigned_exams'),
        total_completed=Count('assigned_exams__exam_sessions', filter=Q(assigned_exams__exam_sessions__is_submitted=True))
        # Note: Score is calculated but not stored in ExamSession model
    )[:10]  # Limit for PDF
    
    # Get recent activity
    recent_sessions = ExamSession.objects.filter(
        is_submitted=True
    ).select_related('user_exam__user', 'user_exam__exam').order_by('-end_time')[:10]
    
    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        spaceBefore=20
    )
    
    # Add title
    title = Paragraph("System Reports & Analytics", title_style)
    elements.append(title)
    
    # Add generation info
    gen_info = Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
    elements.append(gen_info)
    elements.append(Spacer(1, 20))
    
    # Exam Performance Section
    elements.append(Paragraph("Exam Performance", section_style))
    
    exam_data = [['Exam Name', 'Course', 'Assigned', 'Completed', 'Rate', 'Avg Score']]
    for exam in exams:
        completion_rate = (exam.total_completed / exam.total_assigned * 100) if exam.total_assigned > 0 else 0
        # Since score is not stored in ExamSession, we can't calculate average score
        exam_data.append([
            exam.exam_title[:15] + '...' if len(exam.exam_title) > 15 else exam.exam_title,
            exam.course.course_name[:12] + '...' if len(exam.course.course_name) > 12 else exam.course.course_name,
            str(exam.total_assigned or 0),
            str(exam.total_completed or 0),
            f"{completion_rate:.1f}%",
            "N/A"
        ])
    
    exam_table = Table(exam_data, colWidths=[1.5*inch, 1.2*inch, 0.7*inch, 0.7*inch, 0.6*inch, 0.8*inch])
    exam_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(exam_table)
    elements.append(Spacer(1, 20))
    
    # User Performance Section
    elements.append(Paragraph("User Performance (Top 10)", section_style))
    
    user_data = [['Full Name', 'National ID', 'Assigned', 'Completed', 'Rate', 'Avg Score']]
    for user in users:
        completion_rate = (user.total_completed / user.total_assigned * 100) if user.total_assigned > 0 else 0
        # Since score is not stored in ExamSession, we can't calculate average score
        user_data.append([
            user.name[:15] + '...' if len(user.name) > 15 else user.name,
            user.national_id,
            str(user.total_assigned or 0),
            str(user.total_completed or 0),
            f"{completion_rate:.1f}%",
            "N/A"
        ])
    
    user_table = Table(user_data, colWidths=[1.5*inch, 1.0*inch, 0.7*inch, 0.7*inch, 0.6*inch, 0.8*inch])
    user_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(user_table)
    elements.append(Spacer(1, 20))
    
    # Recent Activity Section
    elements.append(Paragraph("Recent Activity", section_style))
    
    activity_data = [['User', 'Exam', 'Score', 'Date', 'Duration']]
    for session in recent_sessions:
        # Calculate duration
        if session.start_time and session.end_time:
            duration = session.end_time - session.start_time
            hours, remainder = divmod(duration.total_seconds(), 3600)
            minutes, _ = divmod(remainder, 60)
            duration_str = f"{int(hours)}h {int(minutes)}m"
        else:
            duration_str = "N/A"
            
        activity_data.append([
            session.user_exam.user.name[:12] + '...' if len(session.user_exam.user.name) > 12 else session.user_exam.user.name,
            session.user_exam.exam.exam_title[:15] + '...' if len(session.user_exam.exam.exam_title) > 15 else session.user_exam.exam.exam_title,
            "N/A",  # Score is calculated but not stored in ExamSession
            session.end_time.strftime('%m/%d %H:%M') if session.end_time else "N/A",
            duration_str
        ])
    
    activity_table = Table(activity_data, colWidths=[1.3*inch, 1.5*inch, 0.7*inch, 0.8*inch, 0.7*inch])
    activity_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(activity_table)
    
    # Build PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reports_{timezone.now().strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    
    return response


@login_required
@user_passes_test(is_admin)
def question_edit_view(request, question_id):
    """Edit an existing question."""
    question = get_object_or_404(Question, question_id=question_id)
    
    if request.method == 'POST':
        exam_id = request.POST.get('exam')
        question_type = request.POST.get('question_type')
        question_text = request.POST.get('question_text')
        option_a = request.POST.get('option_a', '')
        option_b = request.POST.get('option_b', '')
        option_c = request.POST.get('option_c', '')
        option_d = request.POST.get('option_d', '')
        correct_answer = request.POST.get('correct_answer')
        points = request.POST.get('points', 1.0)
        
        try:
            exam = get_object_or_404(Exam, exam_id=exam_id)
            
            # Update question fields
            question.exam = exam
            question.question_type = question_type
            question.question_text = question_text
            question.option_a = option_a
            question.option_b = option_b
            question.option_c = option_c
            question.option_d = option_d
            question.correct_answer = correct_answer
            question.points = float(points) if points else 1.0
            question.save()
            
            messages.success(request, 'Question updated successfully!')
            return redirect('administration:question_list')
            
        except Exception as e:
            messages.error(request, f'Error updating question: {str(e)}')
    
    exams = Exam.objects.all().order_by('exam_title')
    
    context = {
        'question': question,
        'exams': exams,
        'question_types': Question.QUESTION_TYPES,
    }
    
    return render(request, 'administration/question_form.html', context)


@login_required
@user_passes_test(is_admin)
def question_preview_view(request, question_id):
    """Preview a question (AJAX endpoint)."""
    question = get_object_or_404(Question, question_id=question_id)
    
    # Build options HTML for multiple choice
    options_html = ""
    if question.question_type == 'multiple_choice':
        options = [
            ('A', question.option_a),
            ('B', question.option_b),
            ('C', question.option_c),
            ('D', question.option_d)
        ]
        options_html = '<div class="options mb-3"><strong>Options:</strong><ul class="list-unstyled mt-2">'
        for label, option_text in options:
            if option_text:  # Only show non-empty options
                is_correct = question.correct_answer == label
                correct_class = ' class="text-success fw-bold"' if is_correct else ''
                correct_icon = ' <i class="fas fa-check text-success"></i>' if is_correct else ''
                options_html += f'<li{correct_class}><strong>{label}:</strong> {option_text}{correct_icon}</li>'
        options_html += '</ul></div>'
    
    # Build correct answer display
    if question.question_type == 'true_false':
        correct_answer_display = f'<span class="badge bg-success">{"True" if question.correct_answer == "True" else "False"}</span>'
    elif question.question_type == 'multiple_choice':
        correct_answer_display = f'<span class="badge bg-success">Option {question.correct_answer}</span>'
    else:
        correct_answer_display = f'<span class="badge bg-info">Text Answer</span>'
    
    html_content = f"""
    <div class="question-preview">
        <h6>Question Type: {question.get_question_type_display()}</h6>
        <div class="question-text mb-3">
            <strong>Question:</strong> {question.question_text}
        </div>
        
        {options_html}
        
        <div class="correct-answer mb-3">
            <strong>Correct Answer:</strong> 
            {correct_answer_display}
        </div>
        
        <div class="exam-info">
            <small class="text-muted">
                <strong>Exam:</strong> {question.exam.exam_title} - {question.exam.course.course_name}
            </small>
        </div>
    </div>
    """
    
    return JsonResponse({
        'success': True,
        'html': html_content
    })


@login_required
@user_passes_test(is_admin)
def question_delete_view(request, question_id):
    """Delete a question (AJAX endpoint)."""
    if request.method == 'POST':
        try:
            question = get_object_or_404(Question, question_id=question_id)
            question.delete()
            return JsonResponse({
                'success': True,
                'message': 'Question deleted successfully!'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })
